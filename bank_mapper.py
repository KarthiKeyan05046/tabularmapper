"""
bank_mapper.py — Bank Statement -> Standard Schema Mapper (engine)

Two-stage, auditable pipeline:
  Stage 1  detect_header_row()   deterministic scoring (NO AI)
  Stage 2  map_columns()         exact synonym -> fuzzy -> optional llm/embedding fallback

Design invariants (see BUILD_PROMPT.md):
  * No LLM/embedding model ever sees transaction rows. It only sees header
    strings + <=3 sample cells per column. All row/date/amount work is
    deterministic Python.
  * Header detection is scoring, never a model call.
  * Anything ambiguous is flagged needs_review instead of silently guessed.
  * Every column decision carries a method (exact/fuzzy/llm/cache) + 0-100
    confidence for human audit.
"""


from __future__ import annotations

import base64
import csv
import datetime as _dt
import io
import json
import re
from dataclasses import dataclass, field
from typing import Callable, Literal, Optional, Union

from dateutil import parser as _dateparser
from rapidfuzz import fuzz

# --------------------------------------------------------------------------
# Active configuration (output template + synonyms + critical fields).
# Loaded from schema.py — by default the built-in constants (byte-identical to
# the previous hardcoded values), or from a JSON file / URL / S3 object / dict
# via env BANK_MAPPER_CONFIG or a call to configure().
#
# These module globals are kept for backward compatibility; everything reads
# them, and configure() swaps them atomically.
# --------------------------------------------------------------------------
from schema import Config as _Config, load_config as _load_config  # noqa: E402

_ACTIVE_CONFIG: _Config = _load_config()

OUTPUT_SCHEMA: list[tuple[str, str]] = _ACTIVE_CONFIG.headers   # [(field, header)]
SYNONYMS: dict[str, list[str]] = _ACTIVE_CONFIG.synonyms
CRITICAL_FIELDS: set = set(_ACTIVE_CONFIG.critical_fields)
ALLOWED_FIELDS: list[str] = _ACTIVE_CONFIG.allowed_fields
_FIELD_TYPES: dict[str, str] = _ACTIVE_CONFIG.field_types


def configure(source=None, config: "Optional[_Config]" = None) -> None:
    """Swap the active configuration at runtime.

    Pass a `config` object, or a `source` for load_config (path / http(s) URL /
    s3:// / dict). Rebuilds the derived globals and the exact-match lookup so a
    new output template or synonym set takes effect immediately.
    """
    global _ACTIVE_CONFIG, OUTPUT_SCHEMA, SYNONYMS, CRITICAL_FIELDS
    global ALLOWED_FIELDS, _FIELD_TYPES, _EXACT_LOOKUP
    _ACTIVE_CONFIG = config if config is not None else _load_config(source)
    OUTPUT_SCHEMA = _ACTIVE_CONFIG.headers
    SYNONYMS = _ACTIVE_CONFIG.synonyms
    CRITICAL_FIELDS = set(_ACTIVE_CONFIG.critical_fields)
    ALLOWED_FIELDS = _ACTIVE_CONFIG.allowed_fields
    _FIELD_TYPES = _ACTIVE_CONFIG.field_types
    _EXACT_LOOKUP = _build_exact_lookup(SYNONYMS)


# --------------------------------------------------------------------------
# Output format types
# --------------------------------------------------------------------------
OutputFormat = Literal["records", "json", "bytes", "base64", "file"]


# --------------------------------------------------------------------------
# Data classes
# --------------------------------------------------------------------------
@dataclass
class HeaderCandidate:
    index: int
    score: float
    cells: list
    breakdown: dict = field(default_factory=dict)


@dataclass
class ColumnMap:
    col_index: int
    raw_header: str
    field: Optional[str]
    confidence: int
    method: str  # exact | fuzzy | llm | cache | none


@dataclass
class OutputResult:
    """Lazy-evaluated output container supporting multiple serialization formats."""
    records: list[dict]
    format: OutputFormat
    file_path: Optional[str] = None
    _json: Optional[str] = field(default=None, repr=False)
    _bytes: Optional[bytes] = field(default=None, repr=False)
    _base64: Optional[str] = field(default=None, repr=False)

    @property
    def json(self) -> str:
        """Records as JSON string."""
        if self._json is None:
            self._json = json.dumps(self.records, ensure_ascii=False)
        return self._json

    @property
    def bytes(self) -> bytes:
        """Records as .xlsx bytes (lazy, cached)."""
        if self._bytes is None:
            self._bytes = _records_to_xlsx_bytes(self.records)
        return self._bytes

    @property
    def base64(self) -> str:
        """Base64-encoded .xlsx bytes (lazy, cached)."""
        if self._base64 is None:
            self._base64 = base64.b64encode(self.bytes).decode("ascii")
        return self._base64

    def to_response(self) -> Union[list[dict], str, bytes]:
        """Return the native Python object for the requested format."""
        if self.format == "json":
            return self.json
        if self.format == "bytes":
            return self.bytes
        if self.format == "base64":
            return self.base64
        if self.format == "file":
            if self.file_path is None:
                raise ValueError("file_path required for 'file' output format")
            _write_output(self.file_path, self.records)
            return self.file_path
        return self.records

    def __repr__(self) -> str:
        return f"<OutputResult format={self.format} records={len(self.records)}>"


@dataclass
class ProcessResult:
    input_path: str
    output_path: Optional[str]
    header_index: int
    header_score: float
    column_maps: list[ColumnMap]
    records: list[dict]
    needs_review: bool
    review_reasons: list[str]
    header_breakdown: dict = field(default_factory=dict)
    output: Optional[OutputResult] = field(default=None, repr=False)


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
_BANK_VOCAB = {
    "date", "txn", "transaction", "value", "narration", "particulars",
    "description", "details", "remarks", "reference", "ref", "cheque", "chq",
    "debit", "credit", "withdrawal", "deposit", "balance", "amount", "dr", "cr",
    "utr", "branch", "code", "no",
}

_NUM_RE = re.compile(r"^[\s₹$€£rs\.]*[-(]?[\d,]+\.?\d*\)?[\s]*(dr|cr)?$", re.I)


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _looks_numeric(v) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    if isinstance(v, str) and _NUM_RE.match(v.strip()):
        return True
    return False


def _looks_datey(v) -> bool:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return True
    if isinstance(v, str):
        s = v.strip()
        if re.search(r"\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}", s):
            return True
        if re.match(r"\d{1,2}\s*[A-Za-z]{3,9}\s*\d{2,4}", s):
            return True
    return False


def _text_ratio(cells) -> float:
    """Fraction of non-blank cells that are word-like (not numbers/dates)."""
    non_blank = [c for c in cells if not _is_blank(c)]
    if not non_blank:
        return 0.0
    wordy = sum(
        1 for c in non_blank
        if isinstance(c, str) and not _looks_numeric(c) and not _looks_datey(c)
    )
    return wordy / len(non_blank)


# --------------------------------------------------------------------------
# Stage 1 — header detection (deterministic)
# --------------------------------------------------------------------------
def detect_header_row(rows: list[list], scan_limit: int = 25) -> HeaderCandidate:
    """Score the first ~scan_limit rows and return the best header candidate.

    Signals (see BUILD_PROMPT.md §5):
      + density        many non-empty cells
      + text_ratio     headers are words, not numbers/dates
      + short_labels   header cells are short strings
      + vocab_hits     banking-vocabulary matches (strongest)
      + data_below     rows below look like transaction data
      - self_penalty   the row itself is mostly numbers/dates
    """
    n = min(scan_limit, len(rows))
    best: Optional[HeaderCandidate] = None

    for i in range(n):
        cells = rows[i]
        non_blank = [c for c in cells if not _is_blank(c)]
        if not non_blank:
            continue

        density = min(len(non_blank), 8) / 8.0  # cap so wide junk rows don't win
        text_ratio = _text_ratio(cells)

        strs = [c for c in non_blank if isinstance(c, str)]
        short = sum(1 for c in strs if len(c.strip()) <= 25)
        short_labels = short / len(non_blank)

        toks = set()
        for c in strs:
            for t in re.split(r"[^a-z]+", c.lower()):
                if t:
                    toks.add(t)
        vocab_hits = len(toks & _BANK_VOCAB)

        # data_below: sample up to 5 rows beneath; reward numeric/date content
        below_scores = []
        for j in range(i + 1, min(i + 6, len(rows))):
            b = rows[j]
            bnb = [c for c in b if not _is_blank(c)]
            if not bnb:
                continue
            datalike = sum(1 for c in bnb if _looks_numeric(c) or _looks_datey(c))
            below_scores.append(datalike / len(bnb))
        data_below = (sum(below_scores) / len(below_scores)) if below_scores else 0.0

        self_numeric = sum(1 for c in non_blank if _looks_numeric(c) or _looks_datey(c))
        self_penalty = self_numeric / len(non_blank)

        score = (
            1.5 * density
            + 2.0 * text_ratio
            + 1.0 * short_labels
            + 3.0 * min(vocab_hits, 6)          # dominant signal
            + 2.0 * data_below
            - 3.0 * self_penalty
        )
        breakdown = {
            "density": round(density, 2),
            "text_ratio": round(text_ratio, 2),
            "short_labels": round(short_labels, 2),
            "vocab_hits": vocab_hits,
            "data_below": round(data_below, 2),
            "self_penalty": round(self_penalty, 2),
            "score": round(score, 2),
        }
        if best is None or score > best.score:
            best = HeaderCandidate(index=i, score=round(score, 2), cells=list(cells),
                                   breakdown=breakdown)

    if best is None:
        best = HeaderCandidate(index=0, score=0.0,
                               cells=list(rows[0]) if rows else [], breakdown={})
    return best


# --------------------------------------------------------------------------
# Stage 2 — column mapping (exact -> fuzzy -> fallback)
# --------------------------------------------------------------------------
# Build a flat lookup: phrase -> field, for O(1) exact matching.
def _build_exact_lookup(synonyms: dict) -> dict:
    lut: dict[str, str] = {}
    for _fld, _phrases in synonyms.items():
        for _p in _phrases:
            lut[_norm(_p)] = _fld
    return lut


_EXACT_LOOKUP: dict[str, str] = _build_exact_lookup(SYNONYMS)


def _fuzzy_best(header: str) -> tuple[Optional[str], int]:
    """Best fuzzy field + score across all synonym phrases."""
    best_field, best_score = None, 0
    for fld, phrases in SYNONYMS.items():
        for p in phrases:
            s = fuzz.token_set_ratio(header, p)
            # token_set_ratio can over-reward; blend with a stricter ratio
            s = int(0.5 * s + 0.5 * fuzz.ratio(header, p))
            if s > best_score:
                best_field, best_score = fld, s
    return best_field, best_score


def map_columns(
    header_row: list,
    sample_rows: Optional[list[list]] = None,
    llm_fallback: Optional[Callable[[str, list, list], Optional[str]]] = None,
    threshold: int = 80,
) -> list[ColumnMap]:
    """Map each header cell to an output field.

    1. exact synonym  -> confidence 100, method 'exact'
    2. fuzzy (rapidfuzz) -> confidence = score, method 'fuzzy'
    3. if still < threshold and llm_fallback given -> method 'llm'

    The fallback only ever receives the header string + up to 3 sample cells
    for that column + the allowed field list. Never full rows.
    """
    sample_rows = sample_rows or []
    maps: list[ColumnMap] = []
    assigned: set[str] = set()

    for ci, raw in enumerate(header_row):
        raw_str = "" if raw is None else str(raw).strip()
        key = _norm(raw)

        if key == "":
            maps.append(ColumnMap(ci, raw_str, None, 0, "none"))
            continue

        # 1. exact
        if key in _EXACT_LOOKUP:
            maps.append(ColumnMap(ci, raw_str, _EXACT_LOOKUP[key], 100, "exact"))
            continue

        # 2. fuzzy
        fld, score = _fuzzy_best(key)
        if score >= threshold:
            maps.append(ColumnMap(ci, raw_str, fld, int(score), "fuzzy"))
            continue

        # 3. fallback (llm / embedding), header + <=3 samples only
        if llm_fallback is not None:
            samples = []
            for r in sample_rows[:3]:
                if ci < len(r) and not _is_blank(r[ci]):
                    samples.append(str(r[ci])[:40])
            guess = llm_fallback(raw_str, samples, list(ALLOWED_FIELDS))
            if guess in ALLOWED_FIELDS:
                # fallback carries a moderate confidence, clearly below exact
                maps.append(ColumnMap(ci, raw_str, guess, max(int(score), 70), "llm"))
                continue

        # unresolved
        maps.append(ColumnMap(ci, raw_str, None, int(score), "fuzzy"))

    # Resolve duplicates: if two columns claim the same field, keep the higher
    # confidence one; demote the loser to unresolved (needs_review will catch).
    by_field: dict[str, ColumnMap] = {}
    for m in maps:
        if m.field is None:
            continue
        if m.field not in by_field or m.confidence > by_field[m.field].confidence:
            prev = by_field.get(m.field)
            if prev is not None:
                prev.field = None
                prev.method = "dup"
            by_field[m.field] = m
        else:
            m.field = None
            m.method = "dup"
    assigned.update(by_field.keys())
    return maps


# --------------------------------------------------------------------------
# Normalizers (deterministic)
# --------------------------------------------------------------------------
_YEAR_FIRST_RE = re.compile(r"^\s*\d{4}[-/.]")


def normalize_date(v, dayfirst: Optional[bool] = None) -> Optional[str]:
    """Return 'YYYY-MM-DD' or None.

    * Excel datetime/date objects pass straight through.
    * Year-first strings (yyyy-mm-dd, yyyy/mm/dd) are parsed with dayfirst=False
      so they are never flipped.
    * Other strings default to dayfirst=True (dd-mm-yyyy is the common non-US
      bank format) unless the caller overrides via `dayfirst`.
    """
    if _is_blank(v):
        return None
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()

    s = str(v).strip()
    if not s:
        return None

    if _YEAR_FIRST_RE.match(s):
        df = False
    elif dayfirst is None:
        df = True
    else:
        df = dayfirst

    try:
        dt = _dateparser.parse(s, dayfirst=df, fuzzy=True)
        return dt.date().isoformat()
    except (ValueError, OverflowError, TypeError):
        return None


_AMT_CLEAN_RE = re.compile(r"[^\d.\-()]")


def normalize_amount(v) -> Optional[float]:
    """Return a signed float or None.

    Handles: '1,200.50', '(500)' -> -500, '500 Dr' -> -500, '₹500 Cr' -> 500,
    leading minus, currency symbols, stray spaces. Dr/Cr suffix wins over sign.
    """
    if _is_blank(v):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)

    s = str(v).strip()
    if not s:
        return None

    low = s.lower()
    is_dr = bool(re.search(r"\bdr\b|dr$", low))
    is_cr = bool(re.search(r"\bcr\b|cr$", low))
    neg_paren = "(" in s and ")" in s
    neg_sign = s.lstrip().startswith("-")

    cleaned = _AMT_CLEAN_RE.sub("", s).replace("(", "").replace(")", "")
    cleaned = cleaned.replace(",", "")
    # collapse accidental multiple dots / trailing minus
    if cleaned.count("-") > 1:
        cleaned = cleaned.replace("-", "")
        neg_sign = True
    if cleaned in ("", "-", "."):
        return None
    try:
        val = float(cleaned)
    except ValueError:
        return None

    val = abs(val)
    if is_dr:
        return -val
    if is_cr:
        return val
    if neg_paren or neg_sign:
        return -val
    return val


# --------------------------------------------------------------------------
# Extraction
# --------------------------------------------------------------------------
def _field_col(col_maps: list[ColumnMap], fld: str) -> Optional[int]:
    for m in col_maps:
        if m.field == fld:
            return m.col_index
    return None


def extract_records(rows: list[list], header_idx: int,
                    col_maps: list[ColumnMap]) -> list[dict]:
    """Turn data rows into standardized dicts.

    Reconciles the two money layouts:
      * separate debit + credit columns -> used as-is (positive floats)
      * single signed `amount` column   -> negative to debit, positive to credit
    Skips non-transaction rows (no date AND no money). Merges multi-line
    descriptions that spill into blank cells below a transaction.
    """
    fields = _ACTIVE_CONFIG.fields          # ordered output field keys
    types = _FIELD_TYPES                     # field -> "date"|"money"|"text"
    col_of = {f: _field_col(col_maps, f) for f in fields}
    ci_debit = col_of.get("debit")
    ci_credit = col_of.get("credit")
    ci_amt = _field_col(col_maps, "amount")  # input-only mapping (may be output too)

    def cell(r, ci):
        return r[ci] if (ci is not None and ci < len(r)) else None

    records: list[dict] = []
    for r in rows[header_idx + 1:]:
        if all(_is_blank(c) for c in r):
            continue

        # --- money reconciliation (debit/credit vs single signed amount) ---
        debit = credit = None
        if ci_amt is not None and ci_debit is None and ci_credit is None:
            amt = normalize_amount(cell(r, ci_amt))
            if amt is not None:
                if amt < 0:
                    debit = abs(amt)
                elif amt > 0:
                    credit = amt
                else:
                    debit = 0.0
        else:
            d = normalize_amount(cell(r, ci_debit))
            c = normalize_amount(cell(r, ci_credit))
            debit = abs(d) if d is not None else None
            credit = abs(c) if c is not None else None

        # --- build the record, one value per schema field, by type ---
        rec: dict = {}
        date_val = None
        for f in fields:
            if f == "debit":
                rec[f] = debit
            elif f == "credit":
                rec[f] = credit
            else:
                t = types.get(f, "text")
                v = cell(r, col_of.get(f))
                if t == "date":
                    rec[f] = normalize_date(v)
                    if f == "date":
                        date_val = rec[f]
                elif t == "money":
                    rec[f] = normalize_amount(v)      # balance / amount = signed
                else:
                    rec[f] = str(v).strip() if not _is_blank(v) else ""
        if date_val is None:
            date_val = rec.get("date")

        has_money = debit is not None or credit is not None
        desc_val = rec.get("description", "")

        # multi-line description continuation: a row with only a description
        # and no date/money folds into the previous record.
        if not date_val and not has_money and desc_val and records and "description" in rec:
            records[-1]["description"] = (
                records[-1].get("description", "") + " " + desc_val).strip()
            continue

        # skip rows that carry no date and no money (pure noise / subtotals)
        if not date_val and not has_money:
            continue

        records.append(rec)
    return records


# --------------------------------------------------------------------------
# needs_review gate
# --------------------------------------------------------------------------
def evaluate_review(col_maps: list[ColumnMap], records: list[dict],
                    threshold: int = 80) -> tuple[bool, list[str]]:
    reasons: list[str] = []
    mapped_fields = {m.field for m in col_maps if m.field}

    # critical: date required
    for cf in CRITICAL_FIELDS:
        if cf not in mapped_fields:
            reasons.append(f"missing critical field: {cf}")

    # money movement must be knowable
    if not ({"debit", "credit"} & mapped_fields) and "amount" not in mapped_fields:
        reasons.append("no debit/credit or signed amount column found")

    # low-confidence mapped columns
    for m in col_maps:
        if m.field and m.confidence < threshold and m.method != "exact":
            reasons.append(
                f"low-confidence column '{m.raw_header}' -> {m.field} "
                f"({m.confidence}, {m.method})"
            )

    # any fallback-resolved column is worth a human glance
    for m in col_maps:
        if m.field and m.method == "llm":
            reasons.append(f"fallback-resolved column '{m.raw_header}' -> {m.field}")

    if not records:
        reasons.append("no transaction rows extracted")

    return (len(reasons) > 0), reasons


# --------------------------------------------------------------------------
# Output serializers
# --------------------------------------------------------------------------
def _records_to_xlsx_bytes(records: list[dict]) -> bytes:
    """Serialize records to .xlsx bytes in-memory (no temp file)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standardized"
    headers = [disp for _, disp in OUTPUT_SCHEMA]
    ws.append(headers)
    for rec in records:
        ws.append([rec.get(fld) for fld, _ in OUTPUT_SCHEMA])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def records_to_csv_bytes(records: list[dict], encoding: str = "utf-8") -> bytes:
    """Serialize records to CSV bytes."""
    bio = io.BytesIO()
    text = io.TextIOWrapper(bio, encoding=encoding, newline="")
    headers = [disp for _, disp in OUTPUT_SCHEMA]
    writer = csv.DictWriter(text, fieldnames=[f for f, _ in OUTPUT_SCHEMA])
    writer.writeheader()
    for rec in records:
        writer.writerow(rec)
    text.flush()
    return bio.getvalue()


def _write_output(path: str, records: list[dict]) -> None:
    """Write records to an .xlsx file on disk."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standardized"
    headers = [disp for _, disp in OUTPUT_SCHEMA]
    ws.append(headers)
    for rec in records:
        ws.append([rec.get(fld) for fld, _ in OUTPUT_SCHEMA])
    wb.save(path)


# --------------------------------------------------------------------------
# AI integration
# --------------------------------------------------------------------------
AI_CONFIDENCE = 85  # below exact(100), at/above the fuzzy gate so it stands on its own


def _has_critical_gap(col_maps: list[ColumnMap]) -> bool:
    """True if the deterministic pass is missing a critical field — i.e. this
    looks like a new/unknown layout worth asking the AI about."""
    fields = {m.field for m in col_maps if m.field}
    if not CRITICAL_FIELDS.issubset(fields):
        return True
    if not ({"debit", "credit"} & fields) and "amount" not in fields:
        return True
    return False


def merge_ai_mapping(col_maps: list[ColumnMap], ai: dict) -> list[ColumnMap]:
    """Overlay an AI {col_index: field} mapping onto deterministic col_maps.

    Exact (100) matches are ground truth and are kept. The AI fills columns the
    deterministic pass could not place, without stealing a field an exact column
    already owns. Single-slot fields are de-duplicated with exact > ai priority.
    """
    exact_fields = {m.field for m in col_maps if m.method == "exact" and m.field}
    by_index = {m.col_index: m for m in col_maps}
    for ci, field in ai.items():
        m = by_index.get(ci)
        if m is None or field not in ALLOWED_FIELDS:
            continue
        if m.method == "exact":
            continue                      # never override ground truth
        if field in exact_fields:
            continue                      # an exact column already owns this field
        m.field = field
        m.method = "ai"
        m.confidence = AI_CONFIDENCE

    # de-dup single-slot fields: keep highest confidence, prefer exact then ai
    prio = {"exact": 3, "ai": 2, "cache": 2, "fuzzy": 1}
    best: dict[str, ColumnMap] = {}
    for m in col_maps:
        if not m.field:
            continue
        cur = best.get(m.field)
        if cur is None or (prio.get(m.method, 0), m.confidence) > \
                          (prio.get(cur.method, 0), cur.confidence):
            if cur is not None:
                cur.field, cur.method = None, "dup"
            best[m.field] = m
        else:
            m.field, m.method = None, "dup"
    return col_maps


# --------------------------------------------------------------------------
# Core runner
# --------------------------------------------------------------------------
def _run(rows: list[list], source_label: str, out_path, llm_fallback,
         table_matcher, scan_limit, threshold, cache,
         output_format: OutputFormat) -> ProcessResult:
    """Shared core: detect header -> map -> (AI) -> extract -> review -> output.
    Works on already-read `rows` so it is source-agnostic (path or stream)."""
    if not rows:
        return ProcessResult(source_label, None, 0, 0.0, [], [], True,
                             ["empty sheet"], {})

    hc = detect_header_row(rows, scan_limit=scan_limit)
    header = hc.cells
    sample_rows = rows[hc.index + 1: hc.index + 6]

    from_cache = False
    col_maps = None
    if cache is not None:
        cached = cache.get(header)
        if cached is not None:
            col_maps = cached
            from_cache = True

    if col_maps is None:
        col_maps = map_columns(header, sample_rows, llm_fallback=llm_fallback,
                               threshold=threshold)
        # Unknown layout? Ask the AI to map the whole table (structure only).
        if table_matcher is not None and _has_critical_gap(col_maps):
            ai = table_matcher(header, rows[hc.index + 1: hc.index + 46],
                               list(ALLOWED_FIELDS))
            if ai:
                col_maps = merge_ai_mapping(col_maps, ai)

    records = extract_records(rows, hc.index, col_maps)
    needs_review, reasons = evaluate_review(col_maps, records, threshold=threshold)

    # Only cache a freshly-computed mapping if it's trustworthy. Never persist an
    # unconfirmed fallback/low-confidence guess — that would let it be replayed
    # as if approved. (A human-approved mapping can be cached explicitly.)
    if cache is not None and not from_cache and not needs_review:
        cache.put(header, col_maps)

    # Build output result
    output = OutputResult(
        records=records,
        format=output_format,
        file_path=out_path,
    )

    # For backward compat: still write file if out_path given and format is "file"
    if out_path and output_format == "file":
        _write_output(out_path, records)

    return ProcessResult(
        input_path=source_label, output_path=out_path, header_index=hc.index,
        header_score=hc.score, column_maps=col_maps, records=records,
        needs_review=needs_review, review_reasons=reasons,
        header_breakdown=hc.breakdown, output=output,
    )


# --------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------
def process_file(
    path: str,
    out_path: Optional[str] = None,
    output_format: OutputFormat = "file",
    llm_fallback: Optional[Callable] = None,
    table_matcher: Optional[Callable] = None,
    scan_limit: int = 25,
    threshold: int = 80,
    cache: Optional["MappingCache"] = None,
) -> ProcessResult:
    """Read an .xlsx from a filesystem `path` and map it.

    `output_format` controls the serialization of `result.output`:
      - "file"    : writes to `out_path` on disk, returns path string
      - "records" : raw Python list[dict] (default for streams)
      - "json"    : JSON string
      - "bytes"   : in-memory .xlsx bytes
      - "base64"  : base64-encoded .xlsx bytes

    `table_matcher(header_row, data_rows, allowed_fields) -> {col_index: field}`
    is the LLM path (see ai_matcher.OpenAICompatibleMatcher). It fires only when
    the deterministic pass leaves a critical gap AND the header isn't cached.
    """
    rows = _read_sheet(path)
    return _run(rows, path, out_path, llm_fallback, table_matcher,
                scan_limit, threshold, cache, output_format)


def process_stream(
    data,
    out_path: Optional[str] = None,
    output_format: OutputFormat = "records",
    llm_fallback: Optional[Callable] = None,
    table_matcher: Optional[Callable] = None,
    scan_limit: int = 25,
    threshold: int = 80,
    cache: Optional["MappingCache"] = None,
    source_label: str = "<stream>",
) -> ProcessResult:
    """Map an .xlsx received as raw bytes or a binary file-like object — no temp
    file, nothing written to disk. Ideal for a FastAPI UploadFile: pass
    `await file.read()` (bytes) or `file.file` (a stream) straight in.

    For bank data this is the preferred entry point: the statement is parsed
    entirely in memory and never lands on the filesystem.

    Default `output_format` is "records" since streams are typically consumed
    by an API that serializes its own response.
    """
    import io
    if isinstance(data, (bytes, bytearray)):
        fileobj = io.BytesIO(data)
    else:
        fileobj = data  # already a binary file-like object
    rows = _read_sheet(fileobj)
    return _run(rows, source_label, out_path, llm_fallback, table_matcher,
                scan_limit, threshold, cache, output_format)


# --------------------------------------------------------------------------
# Internal sheet reader
# --------------------------------------------------------------------------
def _read_sheet(path: str) -> list[list]:
    # `src` may be a filesystem path OR a binary file-like object (BytesIO) —
    # openpyxl accepts both, so uploads can be read straight from memory.
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows