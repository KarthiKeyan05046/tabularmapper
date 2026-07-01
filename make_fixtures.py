#!/usr/bin/env python3
"""
make_fixtures.py — generate synthetic test statements in test_statements/.

Covers the four required scenarios (BUILD_PROMPT.md §8.4):
  1. junk rows on top + separate Debit/Credit columns
  2. title row + single signed Amount column
  3. header on row 1 + parentheses-for-negative debits (signed amount)
  4. unknown / garbage format that must trip needs_review

Plus a 5th 'weird-header' fixture whose header only resolves via the
embedding/hashing fallback (exercises the fallback path).
"""

import datetime as dt
import os

import openpyxl

HERE = os.path.dirname(__file__)
OUT = os.path.join(HERE, "test_statements")
os.makedirs(OUT, exist_ok=True)


def _save(wb, name):
    wb.save(os.path.join(OUT, name))
    print("wrote", name)


def fixture_junk_split():
    """Junk/metadata rows on top, then Debit/Credit split columns."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["MEGA BANK LTD", None, None, None, None])
    ws.append(["Account Statement", None, None, None, None])
    ws.append(["Account No", "123456789", None, None, None])
    ws.append(["Statement Period", "01-04-2025 to 30-04-2025", None, None, None])
    ws.append([None, None, None, None, None])
    ws.append(["Txn Date", "Narration", "Ref No./Cheque No.", "Withdrawal", "Deposit"])
    ws.append(["01-04-2025", "OPENING BALANCE", "", None, None])
    ws.append(["02-04-2025", "UPI PAYMENT TO STORE", "REF100", "1,299.50", None])
    ws.append(["03-04-2025", "SALARY CREDIT", "REF101", None, "45,000.00"])
    ws.append(["05-04-2025", "ATM WITHDRAWAL", "REF102", "2000", None])
    _save(wb, "01_junk_split.xlsx")


def fixture_title_signed():
    """One title row, then a single signed Amount column."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["HDFC-style export — transactions", None, None, None])
    ws.append(["Date", "Particulars", "Cheque No", "Amount"])
    ws.append(["2025/06/01", "NEFT INWARD", "N001", "3500"])      # yyyy/mm/dd -> credit
    ws.append(["2025/06/02", "CARD SPEND",  "",     "-1299"])     # negative -> debit
    ws.append(["2025/06/03", "IMPS INWARD", "N002", "12,000.75"]) # credit
    ws.append(["2025/06/04", "BILL PAY",    "N003", "500 Dr"])    # Dr -> debit
    _save(wb, "02_title_signed.xlsx")


def fixture_header_row1_brackets():
    """Header on row 1, parentheses-for-negative in a signed amount column."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Transaction Date", "Description", "Reference", "Amount", "Balance"])
    ws.append([dt.datetime(2025, 6, 1), "OPENING", "", "(0)", 10000])
    ws.append([dt.datetime(2025, 6, 2), "GROCERY", "R1", "(299)", 9701])   # -299 debit
    ws.append([dt.datetime(2025, 6, 3), "REFUND",  "R2", "150", 9851])     # +150 credit
    ws.append([dt.datetime(2025, 6, 4), "RENT",    "R3", "(8500)", 1351])  # -8500 debit
    _save(wb, "03_header_row1_brackets.xlsx")


def fixture_garbage():
    """No recognizable transaction table -> must set needs_review."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Zog", "Blorp", "Fizz", "Quux"])
    ws.append([42, 3.14, "lorem", "ipsum"])
    ws.append(["alpha", "beta", "gamma", "delta"])
    ws.append([1, 2, 3, 4])
    _save(wb, "04_garbage.xlsx")


def fixture_weird_header():
    """Money columns miss exact+fuzzy but resolve via the offline fallback.

    'Money out' / 'Money in' are not in SYNONYMS and score below the fuzzy
    threshold, so without a fallback the statement trips needs_review (no
    debit/credit/amount). The lexical HashingEmbeddingFallback recovers them
    because those phrases appear in the field descriptions; the semantic
    MiniLM EmbeddingFallback additionally handles cases with no word overlap.
    """
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["Txn Date", "Narration", "Ref No.", "Outgoing", "Incoming"])
    ws.append(["01-06-2025", "Coffee shop", "D1", "150", None])
    ws.append(["02-06-2025", "Paycheck", "D2", None, "50000"])
    ws.append(["03-06-2025", "Groceries", "D3", "2300", None])
    _save(wb, "05_weird_header.xlsx")


if __name__ == "__main__":
    fixture_junk_split()
    fixture_title_signed()
    fixture_header_row1_brackets()
    fixture_garbage()
    fixture_weird_header()
    print("done ->", OUT)
