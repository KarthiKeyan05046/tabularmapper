"""
Tests for the externalized schema/config system (schema.py + engine.configure).

Guarantees:
  * default config is byte-identical to the legacy hardcoded constants
  * a custom schema (rename / reorder / drop / ADD a new field) drives extraction
  * config loads from a JSON file and from an in-memory dict
  * a bad/unreachable source fails safe to the defaults
Every test that calls configure() restores defaults afterward.
"""

import io
import json
import os
import sys

import pytest

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "src"))

from schema_mapper import engine as bm          # noqa: E402
from schema_mapper import schema                     # noqa: E402

FIX = os.path.join(ROOT, "test_statements")
SAMPLES = os.path.join(ROOT, "samples")


@pytest.fixture(autouse=True)
def _reset_config():
    """Ensure every test starts and ends on the default config."""
    bm.configure()
    yield
    bm.configure()


def test_default_config_is_empty():
    # This is a general mapper — with no config it maps nothing.
    assert bm.OUTPUT_SCHEMA == []
    assert bm.SYNONYMS == {}
    assert bm.CRITICAL_FIELDS == set()


def test_bank_preset_gives_bank_schema():
    bm.configure(config=schema.bank_preset())
    assert bm.OUTPUT_SCHEMA == [
        ("date", "Date"), ("description", "Narration"),
        ("reference", "Reference Number"), ("debit", "Debit"),
        ("credit", "Credit"), ("balance", "Balance"),
    ]
    assert bm.SYNONYMS == schema.BANK_SYNONYMS
    assert bm.CRITICAL_FIELDS == {"date"}


def test_bank_preset_extraction():
    bm.configure(config=schema.bank_preset())
    res = bm.process_file(os.path.join(FIX, "01_junk_split.xlsx"))
    assert res.header_index == 5
    assert res.needs_review is False
    assert list(res.records[0].keys()) == ["date", "description", "reference",
                                           "debit", "credit", "balance"]


def test_custom_schema_rename_reorder_drop():
    cfg = schema.config_from_dict({
        "output_schema": [
            {"field": "date", "header": "When", "type": "date"},
            {"field": "credit", "header": "In", "type": "money"},
            {"field": "debit", "header": "Out", "type": "money"},
        ],
        "synonyms": schema.BANK_SYNONYMS,     # need synonyms to match the headers
    })
    bm.configure(config=cfg)
    assert [h for _, h in bm.OUTPUT_SCHEMA] == ["When", "In", "Out"]
    res = bm.process_file(os.path.join(FIX, "01_junk_split.xlsx"))
    # only the three schema fields appear, in schema order
    assert list(res.records[0].keys()) == ["date", "credit", "debit"]
    assert any(r["debit"] == 1299.5 for r in res.records)
    assert any(r["credit"] == 45000.0 for r in res.records)


def test_custom_schema_adds_new_field_type():
    """A brand-new 'value_date' field is extracted generically by its type."""
    cfg = schema.config_from_dict({
        "output_schema": [
            {"field": "date", "header": "Txn Date", "type": "date"},
            {"field": "value_date", "header": "Value Date", "type": "date"},
            {"field": "description", "header": "Description", "type": "text"},
            {"field": "debit", "header": "Debit", "type": "money"},
            {"field": "credit", "header": "Credit", "type": "money"},
        ],
        "synonyms": dict(schema.BANK_SYNONYMS,
                         date=["txn date"], value_date=["value date"]),
    })
    bm.configure(config=cfg)
    p = os.path.join(SAMPLES, "PAYIR_FC_SBI_2025.xlsx")
    if not os.path.exists(p):
        pytest.skip("sample not present")
    res = bm.process_file(p)
    r0 = res.records[0]
    assert "value_date" in r0 and r0["value_date"]      # new field extracted
    assert r0["date"] and r0["date"] != None            # original date still there


def test_generic_non_engine():
    """The engine has no hardcoded bank fields: a product-catalog config with no
    date/debit/credit, no reconcile and no require_any maps and extracts cleanly
    and is NOT flagged for review."""
    import io
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["ACME price list", None, None])          # junk row on top
    ws.append(["SKU", "Product Name", "Unit Price"])
    ws.append(["A-100", "Widget", "12.50"])
    ws.append(["A-200", "Gadget", "7.99"])
    buf = io.BytesIO(); wb.save(buf)

    cfg = schema.config_from_dict({
        "output_schema": [
            {"field": "sku", "header": "SKU", "type": "text"},
            {"field": "name", "header": "Product Name", "type": "text"},
            {"field": "price", "header": "Unit Price", "type": "number"}],
        "synonyms": {"sku": ["sku"], "name": ["product name"],
                     "price": ["unit price"]},
        "replace_synonyms": True,
        "critical_fields": ["sku"],
    })
    bm.configure(config=cfg)
    res = bm.process_stream(buf.getvalue())
    assert res.header_index == 1
    assert {m.raw_header: m.field for m in res.column_maps if m.field} == {
        "SKU": "sku", "Product Name": "name", "Unit Price": "price"}
    assert res.records == [
        {"sku": "A-100", "name": "Widget", "price": 12.5},
        {"sku": "A-200", "name": "Gadget", "price": 7.99}]
    assert res.needs_review is False          # no bank require_any rule imposed


def test_load_config_from_file(tmp_path):
    cfg_path = tmp_path / "config.json"
    cfg_path.write_text(json.dumps({
        "output_schema": [{"field": "date", "header": "D", "type": "date"},
                          {"field": "amount", "header": "Amt", "type": "money"}],
        "synonyms": {"date": ["date"], "amount": ["amount"]},
    }))
    cfg = schema.load_config(str(cfg_path))
    assert [f.field for f in cfg.output_schema] == ["date", "amount"]
    assert cfg.headers[1] == ("amount", "Amt")


def test_bad_source_fails_safe_to_defaults():
    cfg = schema.load_config("s3://nope/missing.json")     # unreachable
    assert cfg.fields == []                                 # default is now empty
    with pytest.raises(Exception):
        schema.load_config("s3://nope/missing.json", strict=True)


def test_serializers_follow_active_schema():
    cfg = schema.config_from_dict({
        "output_schema": [{"field": "date", "header": "MyDate", "type": "date"},
                          {"field": "debit", "header": "MyDebit", "type": "money"},
                          {"field": "credit", "header": "MyCredit", "type": "money"}],
    })
    bm.configure(config=cfg)
    raw = open(os.path.join(FIX, "01_junk_split.xlsx"), "rb").read()
    res = bm.process_stream(raw)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(res.output.bytes))
    assert [c.value for c in next(wb.active.iter_rows())] == ["MyDate", "MyDebit", "MyCredit"]
    csv_head = bm.records_to_csv_bytes(res.records).split(b"\n")[0].decode()
    assert csv_head.strip() == "date,debit,credit"
